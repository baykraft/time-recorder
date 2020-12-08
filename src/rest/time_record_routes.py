from flask import Blueprint, request, jsonify, make_response
import requests
import json
import datetime
import logging
from flask_cors import CORS
from plugins.models import Session, TimeRecord, BreakTime, FixedTime, User
from typing import List
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from dateutil.relativedelta import relativedelta

module_api = Blueprint('time_records', __name__)
CORS(module_api)
logger = logging.getLogger('flask.app')


@module_api.url_value_preprocessor
def url_value_preprocessor(endpoint, values):
    """
    パラメータで受け取ったトークンに対する認証を実施します。

    :param endpoint: エンドポイント
    :type endpoint: str
    :param values: パスパラメータリスト
    :type values: dict
    :return: 認証OKの場合は無し、認証NGの場合はJSONメッセージ
    :rtype: tuple[Any, int]
    """
    if 'OPTIONS' != request.method:
        token = request.args.get('token')
        r = requests.post('https://slack.com/api/auth.test', {'token': token})
        j = json.loads(r.text)
        if j['ok']:
            values['user'] = j['user_id']
        else:
            return jsonify({'ok': j['ok'], 'message': j['error']}), 401


@module_api.route('/<int:year>/<int:month>/<int:date>', methods=['DELETE'])
def delete(user: str, year: int, month: int, date: int):
    """
    指定年月日の勤怠記録情報を削除します。

    :param user: ユーザID
    :type user: str
    :param year: 年
    :type year: int
    :param month: 月
    :type month: int
    :param date: 日
    :type date: int
    :return: JSON形式のメッセージ
    :rtype: tuple[Any, int]
    """
    session = Session()
    try:
        record: TimeRecord = session.query(TimeRecord).filter(
            TimeRecord.user == user,
            TimeRecord.date == datetime.date(year, month, date)
        ).first()

        if record:
            session.delete(record)
            return jsonify({'ok': True}), 200
        else:
            return jsonify({'ok': False}), 404
    except Exception as e:
        session.rollback()
        logger.error(e, exc_info=True)
    finally:
        if session.is_active:
            session.commit()
        session.close()


@module_api.route('/<int:year>/<int:month>/<int:date>', methods=['PUT'])
def update(user: str, year: int, month: int, date: int):
    """
    指定年月日の勤怠記録情報を更新します。

    :param user: ユーザID
    :type user: str
    :param year: 年
    :type year: int
    :param month: 月
    :type month: int
    :param date: 日
    :type date: int
    :return: JSON形式のメッセージ
    :rtype: tuple[Any, int]
    """
    session = Session()
    try:
        customer = request.json['customer'] if 'customer' in request.json else None
        kind = request.json['kind'] if 'kind' in request.json else None
        start_time = request.json['start_time'] if 'start_time' in request.json else None
        end_time = request.json['end_time'] if 'end_time' in request.json else None
        note = request.json['note'] if 'note' in request.json else None

        filtered: TimeRecord = session.query(TimeRecord).filter(
            TimeRecord.user == user,
            TimeRecord.date == datetime.date(year, month, date)
        ).first()

        if filtered:
            filtered.customer = customer
            filtered.kind = kind
            filtered.start_time = datetime.datetime.strptime(start_time, '%H:%M').time() if start_time else None
            filtered.end_time = datetime.datetime.strptime(end_time, '%H:%M').time() if end_time else None
            filtered.note = note

            date = datetime.date(year=filtered.date.year, month=filtered.date.month, day=filtered.date.day)
            result = __time_record_to_result(session, user, date, filtered)
            return jsonify({'ok': True, 'record': result}), 200
        else:
            return jsonify({'ok': False}), 404
    except Exception as e:
        session.rollback()
        logger.error(e, exc_info=True)
    finally:
        if session.is_active:
            session.commit()
        session.close()


@module_api.route('/<int:year>/<int:month>/<int:date>', methods=['POST'])
def create(user: str, year: int, month: int, date: int):
    """
    指定年月日の勤怠記録情報を登録します。

    :param user: ユーザID
    :type user: str
    :param year: 年
    :type year: int
    :param month: 月
    :type month: int
    :param date: 日
    :type date: int
    :return: JSON形式のメッセージ
    :rtype: tuple[Any, int]
    """
    session = Session()
    try:
        customer = request.json['customer'] if 'customer' in request.json else None
        kind = request.json['kind'] if 'kind' in request.json else None
        start_time = request.json['start_time'] if 'start_time' in request.json else None
        end_time = request.json['end_time'] if 'end_time' in request.json else None
        note = request.json['note'] if 'note' in request.json else None

        filtered: TimeRecord = session.query(TimeRecord).filter(
            TimeRecord.user == user,
            TimeRecord.date == datetime.date(year, month, date)
        ).first()

        if filtered:
            return jsonify({'ok': False}), 409
        else:
            record: TimeRecord = TimeRecord(user, datetime.date(year, month, date))
            record.start_time = datetime.datetime.strptime(start_time, '%H:%M').time() if start_time else None
            record.end_time = datetime.datetime.strptime(end_time, '%H:%M').time() if end_time else None
            record.note = note
            record.customer = customer
            record.kind = kind
            session.add(record)
            session.flush()

            date = datetime.date(year=record.date.year, month=record.date.month, day=record.date.day)
            result = __time_record_to_result(session, user, date, record)
            return jsonify({'ok': True, 'record': result}), 201
    except Exception as e:
        session.rollback()
        logger.error(e, exc_info=True)
    finally:
        if session.is_active:
            session.commit()
        session.close()


@module_api.route('/<int:year>/<int:month>', methods=['GET'])
def records(user: str, year: int, month: int):
    """
    指定年月の勤怠記録情報を取得します。

    :param user: ユーザID
    :type user: str
    :param year: 年
    :type year: int
    :param month: 月
    :type month: int
    :return: 正常時: 勤怠記録情報リスト, 異常時: JSON形式のエラーメッセージ
    :rtype: tuple[Any, int]
    """
    session = Session()
    try:
        results = __get_time_records(session, user, year, month)
        return jsonify({'ok': True, 'records': results}), 200
    except Exception as e:
        session.rollback()
        logger.error(e, exc_info=True)
    finally:
        if session.is_active:
            session.commit()
        session.close()


@module_api.route('/<int:year>/<int:month>/download', methods=['GET'])
def download(user: str, year: int, month: int):
    """
    勤怠記録をExcelファイルでダウンロードします。

    :param user: ユーザID
    :type user: str
    :param year: 年
    :type year: int
    :param month: 月
    :type month: int
    :return: Excelファイル
    :rtype: Any
    """
    session = Session()
    try:
        wb = load_workbook('resources/template.xlsx')
        ws = wb['勤怠']
        ws.cell(row=4, column=14).value = f'{year}年{month}月'

        # ユーザ情報取得
        u: User = session.query(User).filter(
            User.user == user
        ).first()
        if u:
            ws.cell(row=6, column=13).value = u.real_name

        # 日付毎の勤怠記録を生成
        results = __get_time_records(session, user, year, month)

        # 所定時間を取得
        fixed_times: List[FixedTime] = session.query(FixedTime).filter(
            FixedTime.user == user,
            FixedTime.year == year,
            FixedTime.month == month
        ).all()

        # 客先別稼働時間出力
        row_index = 11
        sum_fixed_days = 0              # 所定日数合計
        sum_total_fixed_times = '0:00'  # 総所定時間合計
        sum_actual_days = 0             # 実働日数合計
        sum_actual_times = '0:00'       # 実働時間合計
        for ft in fixed_times:
            break_times: List[BreakTime] = session.query(BreakTime).filter(
                BreakTime.user == user,
                BreakTime.year == year,
                BreakTime.month == month,
                BreakTime.customer == ft.customer
            ).all()

            # 所定日数
            fixed_days = len(list(filter(lambda r: not r['holiday'] and ft.customer == r['customer'], results)))
            sum_fixed_days += fixed_days
            # 所定時間, 総所定時間算出
            fixed_time = __calc_fixed_time(break_times, ft)
            fixed_hour = int(fixed_time.split(':')[0]) * fixed_days
            fixed_minute = int(fixed_time.split(':')[1]) * fixed_days
            h, m = divmod(fixed_minute, 60)
            fixed_hour += h
            fixed_minute = m
            sum_fixed_time = f'{fixed_hour}:{str(fixed_minute).zfill(2)}'
            sum_total_fixed_times = __sum_times([sum_total_fixed_times, sum_fixed_time])
            # 実働日数
            actual_list = list(filter(lambda r: ft.customer == r['customer'], results))
            actual_days = len(actual_list)
            one_day: tuple = 10, 20, 30, 40
            half_day: tuple = 11, 12, 21, 22, 31, 32, 41, 42
            for actual in actual_list:
                kind = actual['kind']
                if kind in one_day:
                    actual_days -= 1
                elif kind in half_day:
                    actual_days -= 0.5
            sum_actual_days += actual_days
            # 実働時間算出
            actual_time = __sum_times(list(
                map(lambda r: r['total_time'], filter(lambda r: ft.customer == r['customer'], results))))
            sum_actual_times = __sum_times([sum_actual_times, actual_time])
            ws.cell(row=row_index, column=2).value = ft.customer
            ws.cell(row=row_index, column=3).value = fixed_time
            ws.cell(row=row_index, column=4).value = fixed_days
            ws.cell(row=row_index, column=5).value = sum_fixed_time
            ws.cell(row=row_index, column=6).value = actual_days
            ws.cell(row=row_index, column=7).value = actual_time
            row_index += 1

        # 客先合計を出力
        ws.cell(row=15, column=4).value = sum_fixed_days
        ws.cell(row=15, column=5).value = sum_total_fixed_times
        ws.cell(row=15, column=6).value = sum_actual_days
        ws.cell(row=15, column=7).value = sum_actual_times

        day_off = len(list(filter(lambda r: r['kind'] == 10, results)))
        day_off += len(list(filter(lambda r: r['kind'] == 11 or r['kind'] == 12, results))) / 2
        sp_day_off = len(list(filter(lambda r: r['kind'] == 30, results)))
        sp_day_off += len(list(filter(lambda r: r['kind'] == 31 or r['kind'] == 32, results))) / 2
        sb_day_off = len(list(filter(lambda r: r['kind'] == 40, results)))
        sb_day_off += len(list(filter(lambda r: r['kind'] == 41 or r['kind'] == 42, results))) / 2
        ab_day_off = len(list(filter(lambda r: r['kind'] == 20, results)))
        ab_day_off += len(list(filter(lambda r: r['kind'] == 21 or r['kind'] == 22, results))) / 2
        holiday_work = len(list(filter(lambda r: r['kind'] == 50, results)))

        ws.cell(row=13, column=10).value = day_off
        ws.cell(row=13, column=11).value = sp_day_off
        ws.cell(row=13, column=12).value = sb_day_off
        ws.cell(row=13, column=13).value = ab_day_off
        ws.cell(row=13, column=14).value = holiday_work

        # 残業時間算出
        over_time = __sum_times(list(
            map(lambda r: r['over_time'], filter(lambda r: not r['statutory_holiday'], results))))
        # 深夜残業時間算出
        midnight_time = __sum_times(list(
            map(lambda r: r['midnight_time'], filter(lambda r: not r['statutory_holiday'], results))))
        # 法休残業時間算出
        statutory_time = __sum_times(list(
            map(lambda r: r['over_time'], filter(lambda r: r['statutory_holiday'] and r['kind'] == 50, results))))
        # 法休深夜残業時間算出
        statutory_midnight_time = __sum_times(list(
            map(lambda r: r['midnight_time'], filter(lambda r: r['statutory_holiday'] and r['kind'] == 50, results))))
        # 控除時間算出
        deduction_time = __sum_times(list(map(lambda r: r['deduction_time'], results)))

        ws.cell(row=15, column=10).value = over_time
        ws.cell(row=15, column=11).value = midnight_time
        ws.cell(row=15, column=12).value = statutory_time
        ws.cell(row=15, column=13).value = statutory_midnight_time
        ws.cell(row=15, column=14).value = __sum_times([
            over_time, midnight_time, statutory_time, statutory_midnight_time])
        ws.cell(row=15, column=15).value = deduction_time

        row_index = 18
        day_of_week = '月', '火', '水', '木', '金', '土', '日'
        kind = {
            '0': '',
            '10': '有休', '11': '有休(AM)', '12': '有休(PM)',
            '20': '欠勤', '21': '欠勤(AM)', '22': '欠勤(PM)',
            '30': '特休', '31': '特休(AM)', '32': '特休(PM)',
            '40': '代休', '41': '代休(AM)', '42': '代休(PM)',
            '50': '休出'
        }
        for result in results:
            ws.cell(row=row_index, column=2).value = result['date']
            ws.cell(row=row_index, column=3).value = day_of_week[result['day']]
            ws.cell(row=row_index, column=4).value = 'H' if result['holiday'] else None
            ws.cell(row=row_index, column=5).value = result['customer']
            ws.cell(row=row_index, column=6).value = kind[str(result['kind'])]
            ws.cell(row=row_index, column=7).value = result['start_time']
            ws.cell(row=row_index, column=8).value = result['end_time']
            ws.cell(row=row_index, column=9).value = result['total_time']
            ws.cell(row=row_index, column=10).value = result['over_time']
            ws.cell(row=row_index, column=11).value = result['midnight_time']
            ws.cell(row=row_index, column=12).value = result['deduction_time']
            ws.cell(row=row_index, column=13).value = result['note']
            row_index += 1

        res = make_response(save_virtual_workbook(wb))
        res.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        res.headers['Content-Disposition'] = 'attachment; filename=sample.xlsx'
        return res
    except Exception as e:
        session.rollback()
        logger.error(e, exc_info=True)
    finally:
        if session.is_active:
            session.commit()
        session.close()


@module_api.route('/<int:year>/<int:month>/summary', methods=['GET'])
def summary(user: str, year: int, month: int):
    """
    勤怠記録の要約情報を取得します。

    :param user: ユーザID
    :type user: str
    :param year: 年
    :type year: int
    :param month: 月
    :type month: int
    :return: 要約情報
    :rtype: tuple[Any, int]
    """
    session = Session()
    try:
        results = __get_time_records(session, user, year, month)

        # 所定時間を取得
        fixed_times: List[FixedTime] = session.query(FixedTime).filter(
            FixedTime.user == user,
            FixedTime.year == year,
            FixedTime.month == month
        ).all()

        records = []
        for ft in fixed_times:
            break_times: List[BreakTime] = session.query(BreakTime).filter(
                BreakTime.user == user,
                BreakTime.year == year,
                BreakTime.month == month,
                BreakTime.customer == ft.customer
            ).all()

            # 客先指定で絞り込む
            filtered = list(filter(lambda r: r['customer'] == ft.customer, results))

            # 所定日数
            fixed_days = len(list(filter(lambda r: not r['holiday'], filtered)))
            # 総所定時間算出
            fixed_time = __calc_fixed_time(break_times, ft)
            fixed_hour = int(fixed_time.split(':')[0]) * fixed_days
            fixed_minute = int(fixed_time.split(':')[1]) * fixed_days
            h, m = divmod(fixed_minute, 60)
            fixed_hour += h
            fixed_minute = m
            sum_fixed_time = f'{fixed_hour}:{str(fixed_minute).zfill(2)}'
            # 実働時間算出
            actual_time = __sum_times(list(map(lambda r: r['total_time'], filtered)))
            # 残業時間算出
            over_time = __sum_times(list(
                map(lambda r: r['over_time'], filter(lambda r: not r['statutory_holiday'], filtered))))
            # 深夜残業時間算出
            midnight_time = __sum_times(list(
                map(lambda r: r['midnight_time'], filter(lambda r: not r['statutory_holiday'], filtered))))
            # 法休残業時間算出
            statutory_time = __sum_times(list(
                map(lambda r: r['over_time'], filter(lambda r: r['statutory_holiday'] and r['kind'] == 50, filtered))))
            # 法休深夜残業時間算出
            statutory_midnight_time = __sum_times(list(
                map(lambda r: r['midnight_time'],
                    filter(lambda r: r['statutory_holiday'] and r['kind'] == 50, filtered))))
            # 控除時間算出
            deduction_time = __sum_times(list(map(lambda r: r['deduction_time'], filtered)))

            records.append({
                'customer': ft.customer,
                'fixed_time': sum_fixed_time,
                'actual_time': actual_time,
                'over_time': over_time,
                'midnight_time': midnight_time,
                'statutory_time': statutory_time,
                'statutory_midnight_time': statutory_midnight_time,
                'deduction_time': deduction_time
            })

        return jsonify({'ok': True, 'records': records}), 200
    except Exception as e:
        session.rollback()
        logger.error(e, exc_info=True)
    finally:
        if session.is_active:
            session.commit()
        session.close()


def __get_time_records(session, user: str, year: int, month: int) -> list:
    """
    勤怠記録の辞書型オブジェクトリストを取得します。

    :param session: DBセッション
    :type session: Any
    :param user: ユーザID
    :type user: str
    :param year: 年
    :type year: int
    :param month: 月
    :type month: int
    :return: 勤怠記録の辞書型オブジェクトリスト
    :rtype: list
    """
    # 勤怠記録を取得
    time_records: List[TimeRecord] = session.query(TimeRecord).filter(
        TimeRecord.user == user,
        TimeRecord.date >= datetime.date(year, month, 1),
        TimeRecord.date < datetime.date(year, month, 1) + relativedelta(months=1)
    ).all()

    # 日付毎の勤怠記録を生成
    results = []
    date = datetime.date(year, month, 1)
    while month == date.month:
        target_record = next(
            filter(lambda r: r.date.year == year and r.date.month == month and r.date.day == date.day,
                   time_records), None)
        results.append(__time_record_to_result(session, user, date, target_record))
        date = date + datetime.timedelta(days=1)

    return results


def __get_holidays(year: int) -> dict:
    """
    祝日リストを取得します。

    :param year: 年
    :type year: int
    :return: 祝日リスト
    :rtype: dict
    """
    holiday_response = requests.get(f'https://holidays-jp.github.io/api/v1/{year}/date.json')
    if 200 == holiday_response.status_code:
        return holiday_response.json()
    else:
        return {}


def __sum_times(times: List[str]) -> str:
    """
    文字列の時間リストを時間として合計し、文字列として返却します。

    :param times: 文字列の時間リスト
    :type times: List[str]
    :return: 合計時間を表すの文字列
    :rtype: str
    """
    hour = 0
    minute = 0
    for time in times:
        if time:
            hour += int(time.split(':')[0])
            minute += int(time.split(':')[1])
    h, m = divmod(minute, 60)
    hour += h
    minute = m
    return f'{hour}:{str(minute).zfill(2)}'


def __calc_fixed_time(break_times: List[BreakTime], fixed_time: FixedTime) -> str:
    """
    所定時間を算出します。

    :param break_times: 休憩時間リスト
    :type break_times: List[BreakTime]
    :param fixed_time: 所定時間情報
    :type fixed_time: FixedTime
    :return: 所定時間
    :rtype: str
    """
    base_date = datetime.date(year=2020, month=1, day=1)
    dt1: datetime = datetime.datetime.combine(base_date, fixed_time.start_time)
    dt2: datetime = datetime.datetime.combine(base_date, fixed_time.end_time)
    seconds: float = (dt2 - dt1).total_seconds()

    # 所定時間から休憩時間を控除
    for break_time in break_times:
        bt1: datetime = datetime.datetime.combine(base_date, break_time.start_time)
        bt2: datetime = datetime.datetime.combine(base_date, break_time.end_time)
        if bt2 > dt1 and bt1 < dt2:
            start: datetime = dt1 if dt1 > bt1 else bt1
            end: datetime = dt2 if dt2 < bt2 else bt2
            seconds = seconds - (end - start).total_seconds()

    # 休憩時間を控除した所定時間をdatetime.time型に変換
    m, s = divmod(seconds, 60)  # 秒を60で割った答えがm(分), 余りがs(秒)
    h, m = divmod(m, 60)  # 分を60で割った答えがh(時), 余りがm(分)
    return f'{int(h)}:{str(int(m)).zfill(2)}'


def __calc_total_time(
        break_times: List[BreakTime], date: datetime.date, start_time: datetime.time, end_time: datetime.time) -> str:
    """
    合計時間を算出します。

    :param break_times: 休憩時間リスト
    :type break_times: List[BreakTime]
    :param date: 日付
    :type date: datetime.date
    :param start_time: 開始時間
    :type start_time: datetime.time
    :param end_time: 終了時間
    :type end_time: datetime.time
    :return: 合計時間
    :rtype: str
    """
    dt1: datetime = datetime.datetime.combine(date, start_time)
    dt2: datetime = datetime.datetime.combine(date, end_time)
    seconds: float = (dt2 - dt1).total_seconds()

    # 稼働時間から休憩時間を控除
    for break_time in break_times:
        bt1: datetime = datetime.datetime.combine(date, break_time.start_time)
        bt2: datetime = datetime.datetime.combine(date, break_time.end_time)
        if bt2 > dt1 and bt1 < dt2:
            start: datetime = dt1 if dt1 > bt1 else bt1
            end: datetime = dt2 if dt2 < bt2 else bt2
            seconds = seconds - (end - start).total_seconds()

    # 休憩時間を控除した稼働時間をdatetime.time型に変換
    m, s = divmod(seconds, 60)  # 秒を60で割った答えがm(分), 余りがs(秒)
    h, m = divmod(m, 60)  # 分を60で割った答えがh(時), 余りがm(分)
    return f'{int(h)}:{str(int(m)).zfill(2)}'


def __calc_over_time(break_times: List[BreakTime], fixed_time: FixedTime, holiday: bool, record: TimeRecord) -> str:
    """
    残業時間を算出します。

    :param break_times: 休憩時間リスト
    :type break_times: List[BreakTime]
    :param fixed_time: 所定時間情報
    :type fixed_time: FixedTime
    :param holiday: 祝日フラグ
    :type holiday: bool
    :param record: TimeRecordエンティティ
    :type record: TimeRecord
    :return: 残業時間
    :rtype: str
    """
    dt2: datetime = datetime.datetime.combine(record.date, record.end_time)
    midnight: datetime = datetime.datetime.combine(record.date, datetime.time(hour=22, minute=0))
    dt2 = dt2 if dt2 < midnight else midnight

    # 所定時間を超える部分の時間を算出
    ft1: datetime = datetime.datetime.combine(record.date, fixed_time.start_time)
    ft2: datetime = datetime.datetime.combine(record.date, fixed_time.end_time)
    # 休日かつ勤休が休出の場合は実働時間すべてを残業時間とする（休憩時間は除く）
    ft2 = ft1 if holiday and record.kind == 50 else ft2
    if ft2 < dt2:
        return __calc_total_time(break_times, record.date, ft2.time(), dt2.time())
    else:
        return '0:00'


def __calc_midnight_time(break_times: List[BreakTime], record: TimeRecord) -> str:
    """
    深夜残業時間を算出します。

    :param break_times: 休憩時間リスト
    :type break_times: List[BreakTime]
    :param record: TimeRecordエンティティ
    :type record: TimeRecord
    :return: 残業時間
    :rtype: str
    """
    dt2: datetime = datetime.datetime.combine(record.date, record.end_time)
    midnight: datetime = datetime.datetime.combine(record.date, datetime.time(hour=22, minute=0))

    if midnight < dt2:
        return __calc_total_time(break_times, record.date, midnight.time(), dt2.time())
    else:
        return '0:00'


def __calc_deduction_time(fixed_time: str, total_time: str, over_time: str, midnight_time: str, kind: int) -> str:
    """
    控除時間を算出します。

    :param fixed_time: 所定時間
    :type fixed_time: str
    :param total_time: 合計時間
    :type total_time: str
    :param over_time: 残業時間
    :type over_time: str
    :param midnight_time: 深夜残業時間
    :type midnight_time: str
    :param kind: 勤休
    :type kind: int
    :return: 控除時間
    :rtype: str
    """
    fixed_minute: int = __to_minute(fixed_time)
    total_minute: int = __to_minute(total_time)
    over_minute: int = __to_minute(over_time)
    midnight_minute: int = __to_minute(midnight_time)
    # 勤休が指定されている場合は控除時間は無し
    if 0 == kind and total_minute - (over_minute + midnight_minute) < fixed_minute:
        deduction_minute = fixed_minute - (total_minute - (over_minute + midnight_minute))
        h, m = divmod(deduction_minute, 60)  # 分を60で割った答えがh(時), 余りがm(分)
        return f'{int(h)}:{str(int(m)).zfill(2)}'
    else:
        return '0:00'


def __time_record_to_result(session, user: str, date: datetime.date, record: TimeRecord) -> dict:
    """
    TimeRecordエンティティを辞書型オブジェクトに変換します。

    :param session: DBセッション
    :type session: Any
    :param user: ユーザID
    :type user: str
    :param date: 日付
    :type date: datetime.date
    :param record: TimeRecordエンティティ
    :type record: TimeRecord
    :return: 勤怠記録情報を表す辞書型オブジェクト
    :rtype: dict
    """
    fixed_time = '0:00'       # 所定時間
    total_time = '0:00'       # 合計時間
    over_time = '0:00'        # 残業時間
    midnight_time = '0:00'    # 深夜残業時間
    deduction_time = '0:00'   # 控除時間

    japanese_holidays = __get_holidays(date.year)
    day = date.weekday()  # 0:月～6:日
    holiday_key = '{0:%Y-%m-%d}'.format(date)
    holiday = (day == 5 or day == 6 or holiday_key in japanese_holidays)
    statutory_holiday = (day == 6 or holiday_key in japanese_holidays)
    holiday_note = japanese_holidays[holiday_key] if holiday_key in japanese_holidays else None

    if record and record.start_time and record.end_time:

        # 指定された客先の休憩時間リストを取得
        break_times: List[BreakTime] = session.query(BreakTime).filter(
            BreakTime.user == user,
            BreakTime.year == record.date.year,
            BreakTime.month == record.date.month,
            BreakTime.customer == record.customer
        ).all()

        # 指定された客先の所定時間リストを取得
        fixed_times: List[FixedTime] = session.query(FixedTime).filter(
            FixedTime.user == user,
            FixedTime.year == record.date.year,
            FixedTime.month == record.date.month,
            FixedTime.customer == record.customer
        ).all()

        if 0 < len(fixed_times):
            fixed_time = __calc_fixed_time(break_times, fixed_times[0])
            over_time = __calc_over_time(break_times, fixed_times[0], holiday, record)
        total_time = __calc_total_time(break_times, record.date, record.start_time, record.end_time)
        midnight_time = __calc_midnight_time(break_times, record)
        deduction_time = __calc_deduction_time(fixed_time, total_time, over_time, midnight_time, record.kind)

    return {
        'time_record_id': record.time_record_id if record else 0,
        'year': record.date.year if record else date.year,
        'month': record.date.month if record else date.month,
        'date': record.date.day if record else date.day,
        'day': date.weekday(),
        'holiday': holiday,
        'statutory_holiday': statutory_holiday,
        'customer': record.customer if record else None,
        'kind': record.kind if record else 0,
        'start_time': '{0:%H:%M}'.format(record.start_time) if record and record.start_time else None,
        'end_time': '{0:%H:%M}'.format(record.end_time) if record and record.end_time else None,
        'fixed_time': fixed_time,
        'total_time': total_time,
        'over_time': over_time,
        'midnight_time': midnight_time,
        'deduction_time': deduction_time,
        'note': record.note if record else holiday_note
    }


def __to_minute(time: str) -> int:
    """
    時間文字列を分に変換します。

    :param time: 時間文字列
    :type time: str
    :return: 分
    :rtype: int
    """
    if time:
        return int(time.split(':')[0]) * 60 + int(time.split(':')[1])
    else:
        return 0
